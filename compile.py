import openpyxl
import re
import datetime
import calendar
import json

# open an excel workbook and return an openpyxl workbook instance
def open_wb(name):
    try:
        wb = openpyxl.load_workbook(name)
        print('Successfully loaded input worksheet')
        return wb
    except Exception as e:
        print("Exception in open_wb", e)
        return

def create_worksheet(wb, sheet_name):
    ws = wb.create_sheet(sheet_name)
    return ws

def create_workbook(name):
    wb = openpyxl.Workbook()
    wb.save(f'{name}.xlsx')

def list_components(str):
    if str == "":
        return None
    raw = str.split(',')
    new = []
    for component in raw:
        try:
            match = re.match(r'(.*?)(?=\ -)', component).group(1).strip()
            if match != "CONVERTING COST" and match != "SLITTING COST":
                new.append(match)
        except Exception as E:
            # print(E, component, match)
            continue
    return new

def check_if_late(data):
    return data["post_date"] - data["due_date"] > datetime.timedelta(days=0)

def calc_late_duration(data):
    return (data["post_date"] - data["due_date"]).days

def collect_data(wb):
    ws = wb.active

    all_data = {
        "raw": [],
        "years_seen": []
    }

    for i in range(1, ws.max_row + 1):
        if ws['I' + str(i)].value != "Record Production":
            continue
        if ws['E' + str(i)].value != "Posted":
            continue
        if ws['D' + str(i)].value == None:
            continue

        wo_num = ws['A' + str(i)].value
        row_data = {
            "wo_num": ws['A' + str(i)].value,
            "part_num": ws['B' + str(i)].value,
            "part_desc": ws['C' + str(i)].value,
            "type": ws['D' + str(i)].value,
            "status": ws['E' + str(i)].value,
            "post_date": ws['F' + str(i)].value,
            "due_date": ws['G' + str(i)].value,
            "qty": ws['H' + str(i)].value,
            "components": list_components(ws['J' + str(i)].value),
        }

        row_data["late_duration"] = calc_late_duration(row_data)

        if check_if_late(row_data):
            row_data["is_late"] = True

        else:
            row_data["is_late"] = False

        year = row_data["post_date"].year
        if year not in all_data["years_seen"]:
            all_data["years_seen"].append(year)
        
        all_data["raw"].append(row_data)

    all_data["first_date_seen"] = sorted([wo["post_date"] for wo in all_data["raw"]])[0]
    all_data["last_date_seen"] = sorted([wo["post_date"] for wo in all_data["raw"]])[-1]
    
    return all_data

def contains(word, str):
    try:
        return word.lower() in str.lower()
    except Exception as e:
        print (word, str)

def filter_data(data, converting_type=None, lates_only=False, year=None, month=None):
    
    if year != None:
        data = list(filter(lambda wo: wo["post_date"].year == year, data))

    if month != None:
        data = list(filter(lambda wo: wo["post_date"].month == month, data))

    if lates_only:
        data = list(filter(lambda wo: wo["is_late"] == True, data))

    if converting_type != None:
        data = list(filter(lambda wo: contains(converting_type, wo["type"]) == True, data))

    return data
    
def analyze_qty(data):
    qtys = sorted([wo["qty"] for wo in data])
    
    if len(qtys) <= 0:
        return 

    return {
        "wo_count": len(qtys),
        "sum": sum(qtys),
        "avg": round(sum(qtys)/len(qtys),0),
        "median": qtys[int(len(qtys)/2)]
    }

def analyze_late_duration(data):
    late_durations = sorted([wo["late_duration"] for wo in data])
    
    if len(late_durations) <= 0:
        return 

    return {
        "wo_count": len(late_durations),
        "sum": sum(late_durations),
        "avg": round(sum(late_durations)/len(late_durations),0),
        "median": late_durations[int(len(late_durations)/2)]
    }

# groups all the WOs between the previous month and the # of months preceeding
def rolling_range(data, months):
    # make a date range from the end of last month to 3 months ago
    end = add_months(datetime.datetime.today(), -1) 
    # go to the last day of the month
    end = end.replace(day=calendar.monthrange(end.year, end.month)[1])
    start = add_months(end, -(months-1))

    rolling = [wo for wo in data if wo["post_date"] >= start and wo["post_date"] <= end]

    return rolling


# get the stats just for the previous month and compare against the rolling average # of months
def analyze_last_month(data, rolling_duration):
    wo_list = data["raw"]
    stats = {
        "total": {},
        "slit": {},
        "convert": {},
    }

    # gather all WOs from the month previous to todays date
    last_month = add_months(datetime.datetime.today(),-1) 
    last_mo_data = filter_data(wo_list, year=last_month.year, month=last_month.month)

    # gather all WOs from all the months included in the rolling_duration
    rolling_data = rolling_range(wo_list, rolling_duration)
    
    for converting_type in stats.keys():
        stats[converting_type] = {}
        for period in [f"{calendar.month_name[last_month.month]} {last_month.year}", f"Rolling_{rolling_duration}mo"]:
            # use the last month WOs by default
            period_data = last_mo_data

            # switch to the rolling WOs for rolling stats
            if period == f"Rolling_{rolling_duration}mo":
                period_data = rolling_data
            
            qty_stats = analyze_qty(filter_data(period_data, converting_type=converting_type if converting_type != "total" else None, lates_only=False, year=None, month=None))

            late_qty_stats = analyze_qty(filter_data(period_data, converting_type=converting_type if converting_type != "total" else None, lates_only=True, year=None, month=None))

            late_duration_stats = analyze_late_duration(filter_data(period_data, converting_type=converting_type if converting_type != "total" else None, lates_only=True, year=None, month=None))
            
            stats[converting_type][period] = {
                "WO Count": qty_stats["wo_count"] if qty_stats else 0,
                "Total Qty": qty_stats["sum"] if qty_stats else 0,
                "Avg Qty": qty_stats["avg"] if qty_stats else 0,
                "Late WO Count": late_qty_stats["wo_count"] if late_qty_stats else 0,
                "Late Avg Qty": late_qty_stats["avg"] if late_qty_stats else 0,
                "Late Avg Duration": late_duration_stats["avg"] if late_duration_stats else 0,
            }

            if period == f"Rolling_{rolling_duration}mo":
                stats[converting_type][period]["WO Count"] = round(qty_stats["wo_count"] / rolling_duration) if qty_stats else 0

                stats[converting_type][period]["Total Qty"] = round(qty_stats["sum"] / rolling_duration) if qty_stats else 0

                stats[converting_type][period]["Late WO Count"] = round(late_qty_stats["wo_count"] / rolling_duration) if late_qty_stats else 0

        stats[converting_type]["%_change"] = {
            "WO Count": calc_percent_change(stats[converting_type][f"{calendar.month_name[last_month.month]} {last_month.year}"]["WO Count"], stats[converting_type][f"Rolling_{rolling_duration}mo"]["WO Count"]),
            "Total Qty": calc_percent_change(stats[converting_type][f"{calendar.month_name[last_month.month]} {last_month.year}"]["Total Qty"], stats[converting_type][f"Rolling_{rolling_duration}mo"]["Total Qty"]),
            "Avg Qty": calc_percent_change(stats[converting_type][f"{calendar.month_name[last_month.month]} {last_month.year}"]["Avg Qty"], stats[converting_type][f"Rolling_{rolling_duration}mo"]["Avg Qty"]),
            "Late WO Count": calc_percent_change(stats[converting_type][f"{calendar.month_name[last_month.month]} {last_month.year}"]["Late WO Count"], stats[converting_type][f"Rolling_{rolling_duration}mo"]["Late WO Count"]),
            "Late Avg Qty": calc_percent_change(stats[converting_type][f"{calendar.month_name[last_month.month]} {last_month.year}"]["Late Avg Qty"], stats[converting_type][f"Rolling_{rolling_duration}mo"]["Late Avg Qty"]),
            "Late Avg Duration": calc_percent_change(stats[converting_type][f"{calendar.month_name[last_month.month]} {last_month.year}"]["Late Avg Duration"], stats[converting_type][f"Rolling_{rolling_duration}mo"]["Late Avg Duration"]),
        }

    return stats

def calc_percent_change(a, b):
    if b == 0:
        return None
    return round((a - b) / b, 3)

def summarize_late_components(data):
    stats = {}
    raw_data = data["raw"]

    for year in data["years_seen"]:
        stats[year] = {}
        stats[year]["months"] = {}
        seen_components_yearly = {}

        for num in range(1, 13):
            # if the month and year are NOT within the date range then continue
            if not within_date_range(year, num, data["first_date_seen"].date(), data["last_date_seen"].date()):
                continue

            month = calendar.month_name[num]
            filtered_data = filter_data(raw_data, converting_type=None, lates_only=True, year=year, month=num)

            seen_components_monthly = {}
            for wo in filtered_data:
                for component in wo["components"]:
                    if component in seen_components_monthly:
                        seen_components_monthly[component] += 1
                    else:
                        seen_components_monthly[component] = 1
                    if component in seen_components_yearly:
                        seen_components_yearly[component] += 1
                    else:
                        seen_components_yearly[component] = 1

            stats[year]["months"][month] = dict(sorted(seen_components_monthly.items(), key=lambda item: item[1], reverse=True))
        
        stats[year]["components"] = dict(sorted(seen_components_yearly.items(), key=lambda item: item[1], reverse=True))

    return stats

# add some number of months to a given datetime.datetime instance
def add_months(sourcedate, months):
    month = sourcedate.month - 1 + months
    year = sourcedate.year + month // 12
    month = month % 12 + 1
    new_date = datetime.datetime(year, month, 1)

    return new_date

# checks to see if a given year,month is within the date range of the data set
def within_date_range(year, month, start, end):
    # find date range
    
    date_range_start = datetime.date(start.year, start.month, 1)
    date_range_end = datetime.date(end.year, end.month, 1)

    # set date based on passed in year, month
    current_date = datetime.date(year, month, 1)

    # compare
    if current_date < date_range_start or current_date > date_range_end:
        return False
    return True

# summarize the stats for all workorders
def summarize(data):
    stats = {}
    raw_data = data["raw"]

    for year in data["years_seen"]:
        stats[year] = {
            "wo_count": len(filter_data(raw_data, None, False, year, None)),
            "qtys": analyze_qty(filter_data(raw_data, None, False, year, None)),
            "late_count": len(filter_data(raw_data, None, True, year, None)),
            "late_qtys": analyze_qty(filter_data(raw_data, None, True, year, None)),
            "late_durations": analyze_late_duration(filter_data(raw_data, None, True, year, None)),
        }

        for converting_type in ["slit", "convert"]:
            stats[year][converting_type] = {
                "wo_count": len(filter_data(raw_data, converting_type, False, year, None)),
                "qtys": analyze_qty(filter_data(raw_data, converting_type, False, year, None)),
                "late_count": len(filter_data(raw_data, converting_type, True, year, None)),
                "late_qtys": analyze_qty(filter_data(raw_data, converting_type, True, year, None)),
                "late_durations": analyze_late_duration(filter_data(raw_data, converting_type, True, year, None)),
                "months": {},
            }

            for num in range(1, 13):
                # if the month and year are NOT within the date range then continue
                if not within_date_range(year, num, data["first_date_seen"].date(), data["last_date_seen"].date()):
                    continue

                month = calendar.month_name[num]
                stats[year][converting_type]["months"][month] = {
                    "month": month,
                    "month_num": num,
                    "wo_count": len(filter_data(raw_data, converting_type, False, year, num)),
                    "qtys": analyze_qty(filter_data(raw_data, converting_type, False, year, num)),
                    "late_count": len(filter_data(raw_data, converting_type, True, year, num)),
                    "late_qtys": analyze_qty(filter_data(raw_data, converting_type, True, year, num)),
                    "late_durations": analyze_late_duration(filter_data(raw_data, converting_type, True, year, num)),
                }

    return stats

def print_to_json(data, name):
    data_json = json.dumps(data, indent=4, default=str)

    with open(f'{name}.json', 'w', encoding='utf-8') as f:
        f.write(data_json)

def print_excel_results(wb, results):

    ws = create_worksheet(wb, "Results")
    wb.remove(ws)
    ws = create_worksheet(wb, "Results")

    # print out monthly summary
    CONVERTING_TYPES = [
        "slit",
        "convert"
    ]

    HEADINGS = [
        "Slit WO Count",
        "Slit Qty",
        "Slit Avg Qty",
        "Slit Median Qty",
        "Late Slit Count",
        "Late Slit Qty",
        "Late Slit Avg Qty",
        "Late Slit Median Qty",
        "Late Slit Avg Duration",
        "Late Slit Median Duration",
        "Late Slit Ratio (%)",

        "Convert WO Count",
        "Convert Qty",
        "Convert Avg Qty",
        "Convert Median Qty",
        "Late Convert Count",
        "Late Convert Qty",
        "Late Convert Avg Qty",
        "Late Convert Median Qty",
        "Late Convert Avg Duration",
        "Late Convert Median Duration",
        "Late Convert Ratio (%)",
    ]

    for idx, heading in enumerate(HEADINGS):
        ws.cell(row=idx+2, column=1).value = heading

    col_count = 1
    for year in results:
        for month in results[year]["slit"]["months"]:
            ws.cell(row=1, column=col_count+1).value = f'{month}-{year}'

            skip_count = len(HEADINGS)/len(CONVERTING_TYPES)
            for idx, converting_type in enumerate(CONVERTING_TYPES):
                prop = results[year][converting_type]["months"][month]
                
                ws.cell(row=idx*skip_count + 2, column=col_count+1).value = prop["qtys"]["wo_count"] if prop["qtys"] else 0
                ws.cell(row=idx*skip_count + 3, column=col_count+1).value = prop["qtys"]["sum"] if prop["qtys"] else 0
                ws.cell(row=idx*skip_count + 4, column=col_count+1).value = prop["qtys"]["avg"] if prop["qtys"] else 0
                ws.cell(row=idx*skip_count + 5, column=col_count+1).value = prop["qtys"]["median"] if prop["qtys"] else 0

                ws.cell(row=idx*skip_count + 6, column=col_count+1).value = prop["late_qtys"]["wo_count"] if prop["late_qtys"] else 0
                ws.cell(row=idx*skip_count + 7, column=col_count+1).value = prop["late_qtys"]["sum"] if prop["late_qtys"] else 0
                ws.cell(row=idx*skip_count + 8, column=col_count+1).value = prop["late_qtys"]["avg"] if prop["late_qtys"] else 0
                ws.cell(row=idx*skip_count + 9, column=col_count+1).value = prop["late_qtys"]["median"] if prop["late_qtys"] else 0
                
                ws.cell(row=idx*skip_count + 10, column=col_count+1).value = prop["late_durations"]["avg"] if prop["late_durations"] else 0
                ws.cell(row=idx*skip_count + 11, column=col_count+1).value = prop["late_durations"]["median"] if prop["late_durations"] else 0

                if prop["qtys"] and prop["late_qtys"]:
                    late_ratio = round((prop["late_qtys"]["wo_count"] / prop["qtys"]["wo_count"])*100)
                    ws.cell(row=idx*skip_count + 12, column=col_count+1).value = late_ratio
                else:
                    ws.cell(row=idx*skip_count + 12, column=col_count+1).value = 0

            col_count+=1

    return wb

def print_excel_annual_summaries(wb, results):
    ws = wb["Results"]
    START_ROW = ws.max_row + 2

    HEADINGS = [
        "Slit WO Count",
        "Slit Qty",
        "Slit Avg Qty",
        "Slit Median Qty",
        "Late Slit Count",
        "Late Slit Qty",
        "Late Slit Avg Qty",
        "Late Slit Median Qty",
        "Late Slit Avg Duration",
        "Late Slit Median Duration",
        "Late Slit Ratio (%)",

        "Convert WO Count",
        "Convert Qty",
        "Convert Avg Qty",
        "Convert Median Qty",
        "Late Convert Count",
        "Late Convert Qty",
        "Late Convert Avg Qty",
        "Late Convert Median Qty",
        "Late Convert Avg Duration",
        "Late Convert Median Duration",
        "Late Convert Ratio (%)",
    ]
    
    # print out annual summary
    ANNUAL_HEADINGS = [
        "WO Count",
        "Total Qty",
        "Avg Qty",
        "Median Qty",

        "Late WO Count",
        "Late Qty",
        "Late Avg Qty",
        "Late Median Qty",
        "Late Avg Duration",
        "Late Median Duration",

        "Slit WO Count",
        "Slit Qty",
        "Slit Avg Qty",
        "Slit Median Qty",

        "Late Slit WO Count",
        "Late Slit Qty",
        "Late Slit Avg Qty",
        "Late Slit Median Qty",
        "Late Slit Avg Duration",
        "Late Slit Median Duration",
       
        "Convert WO Count",
        "Convert Qty",
        "Convert Avg Qty",
        "Convert Median Qty",

        "Late Convert WO Count",
        "Late Convert Qty",
        "Late Convert Avg Qty",
        "Late Convert Median Qty",
        "Late Convert Avg Duration",
        "Late Convert Median Duration",
    ]

    # print headings
    for idx, heading in enumerate(ANNUAL_HEADINGS):
        ws.cell(row=START_ROW+idx+1, column=1).value = heading

    # print values
    col_count = 1
    for year in results:
        # year
        ws.cell(row=START_ROW, column=col_count+1).value = year
        # "WO Count",
        ws.cell(row=START_ROW+1, column=col_count+1).value = results[year]["wo_count"]
        # "Total Qty,
        ws.cell(row=START_ROW+2, column=col_count+1).value = results[year]["qtys"]["sum"]
        # "Avg Qty",
        ws.cell(row=START_ROW+3, column=col_count+1).value = results[year]["qtys"]["avg"]
        # "Median Qty",
        ws.cell(row=START_ROW+4, column=col_count+1).value = results[year]["qtys"]["median"]
        
        if results[year]["late_qtys"]:
            # "Late WO Count",
            ws.cell(row=START_ROW+5, column=col_count+1).value = results[year]["late_qtys"]["wo_count"]
            # "Late Qty",
            ws.cell(row=START_ROW+6, column=col_count+1).value = results[year]["late_qtys"]["sum"]
            # "Late Avg Qty",
            ws.cell(row=START_ROW+7, column=col_count+1).value = results[year]["late_qtys"]["avg"]
            # "Late Median Qty",
            ws.cell(row=START_ROW+8, column=col_count+1).value = results[year]["late_qtys"]["median"]

        if results[year]["late_durations"]:
            # "Late Avg Duration",
            ws.cell(row=START_ROW+9, column=col_count+1).value = results[year]["late_durations"]["avg"]
            # "Late Median Duration",
            ws.cell(row=START_ROW+10, column=col_count+1).value = results[year]["late_durations"]["median"]
        
        # "Slit WO Count",
        ws.cell(row=START_ROW+11, column=col_count+1).value = results[year]["slit"]["qtys"]["wo_count"]
        # "Slit Qty",
        ws.cell(row=START_ROW+12, column=col_count+1).value = results[year]["slit"]["qtys"]["sum"]
        # "Slit Avg Qty",
        ws.cell(row=START_ROW+13, column=col_count+1).value = results[year]["slit"]["qtys"]["avg"]
        # "Slit Median Qty",
        ws.cell(row=START_ROW+14, column=col_count+1).value = results[year]["slit"]["qtys"]["median"]
        
        if results[year]["slit"]["late_qtys"]:
            # "Late WO Count",
            ws.cell(row=START_ROW+15, column=col_count+1).value = results[year]["slit"]["late_qtys"]["wo_count"]
            # "Late Qty",
            ws.cell(row=START_ROW+16, column=col_count+1).value = results[year]["slit"]["late_qtys"]["sum"]
            # "Late Avg Qty",
            ws.cell(row=START_ROW+17, column=col_count+1).value = results[year]["slit"]["late_qtys"]["avg"]
            # "Late Median Qty",
            ws.cell(row=START_ROW+18, column=col_count+1).value = results[year]["slit"]["late_qtys"]["median"]

        if results[year]["slit"]["late_durations"]:
            # "Late Avg Duration",
            ws.cell(row=START_ROW+19, column=col_count+1).value = results[year]["slit"]["late_durations"]["avg"]
            # "Late Median Duration",
            ws.cell(row=START_ROW+20, column=col_count+1).value = results[year]["slit"]["late_durations"]["median"]
       
        if results[year]["convert"]["qtys"]:
            # "Convert WO Count",
            ws.cell(row=START_ROW+21, column=col_count+1).value = results[year]["convert"]["qtys"]["wo_count"]
            # "Convert Qty",
            ws.cell(row=START_ROW+22, column=col_count+1).value = results[year]["convert"]["qtys"]["sum"]
            # "Convert Avg Qty",
            ws.cell(row=START_ROW+23, column=col_count+1).value = results[year]["convert"]["qtys"]["avg"]
            # "Convert Median Qty",
            ws.cell(row=START_ROW+24, column=col_count+1).value = results[year]["convert"]["qtys"]["median"]

        if results[year]["convert"]["late_qtys"]:
            # "Late WO Count",
            ws.cell(row=START_ROW+25, column=col_count+1).value = results[year]["convert"]["late_qtys"]["wo_count"]
            # "Late Qty",
            ws.cell(row=START_ROW+26, column=col_count+1).value = results[year]["convert"]["late_qtys"]["sum"]
            # "Late Avg Qty",
            ws.cell(row=START_ROW+27, column=col_count+1).value = results[year]["convert"]["late_qtys"]["avg"]
            # "Late Median Qty",
            ws.cell(row=START_ROW+28, column=col_count+1).value = results[year]["convert"]["late_qtys"]["median"]

        if results[year]["convert"]["late_durations"]:
            # "Late Avg Duration",
            ws.cell(row=START_ROW+29, column=col_count+1).value = results[year]["convert"]["late_durations"]["avg"]
            # "Late Median Duration",
            ws.cell(row=START_ROW+30, column=col_count+1).value = results[year]["convert"]["late_durations"]["median"]
        
        col_count += 1

    return wb

# print the top 3 components with late work orders associated with them
def print_excel_components(wb, components):
    ws = wb["Results"]
    START_ROW = ws.max_row + 2
    col_count = 2

    # side headings
    ws.cell(row=START_ROW+1, column=1).value = "Component 1"
    ws.cell(row=START_ROW+2, column=1).value = "Component 1 Count"
    ws.cell(row=START_ROW+3, column=1).value = "Component 2"
    ws.cell(row=START_ROW+4, column=1).value = "Component 2 Count"
    ws.cell(row=START_ROW+5, column=1).value = "Component 3"
    ws.cell(row=START_ROW+6, column=1).value = "Component 3 Count"

    for year in components:
        for month in components[year]["months"]:
            # print month-year
            ws.cell(row=START_ROW, column=col_count).value = f'{month}-{year}'
            
            # print top components
            component_stats = list(components[year]["months"][month].items())
            for idx, (name, count) in enumerate(component_stats):
                if idx >= 3:
                    break
                ws.cell(row=START_ROW+1+idx*2, column=col_count).value = name
                ws.cell(row=START_ROW+2+idx*2, column=col_count).value = count

            col_count += 1

    return wb

def print_excel_last_month(wb, last_month_results):
    ws = wb["Results"]
    START_ROW = ws.max_row + 2
    col_count = 1

    # top headings
    for idx, period in enumerate(last_month_results["total"].keys()):
        ws.cell(row=START_ROW, column=col_count+1+idx).value = period

    # stats
    for i, converting_type in enumerate(last_month_results.keys()):
        for j, period in enumerate(last_month_results[converting_type].keys()):
            for k, stat in enumerate(last_month_results[converting_type][period].keys()):
                # side headings
                ws.cell(row=START_ROW+1+(k+(i*6)), column=col_count).value = f'{converting_type} {stat}'

                # data summary
                ws.cell(row=START_ROW+1+(k+(i*6)), column=col_count+j+1).value = last_month_results[converting_type][period][stat]
    
    return wb

def save_workbook(wb, name):
    wb.save(f"U:\Josh\JD Working Folder\Adheco General\Warehouse\Converting Analysis/{name}_{datetime.datetime.today().strftime('%d%b%Y')}.xlsx")

def console_log_json(data):
    print(json.dumps(data, indent=2, default=str))

def main(wb_name):
    wb = open_wb(wb_name)
    data = collect_data(wb)
    results = summarize(data)
    components = summarize_late_components(data)
    last_month_results = analyze_last_month(data, 3)
    print_to_json(data, "data")
    print_to_json(results, "results")
    print_to_json(components, "components")
    print_to_json(last_month_results, "last_month")
    print_excel_results(wb, results)
    print_excel_components(wb, components)
    print_excel_annual_summaries(wb, results)
    print_excel_last_month(wb, last_month_results)
    save_workbook(wb, "Workorder Analysis")

main('~CRF096_December2024.xlsx')

